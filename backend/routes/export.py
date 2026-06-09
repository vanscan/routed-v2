"""Stops → Excel (.xlsx) export endpoint.

    GET /stops/export/xlsx  → all stops as a styled workbook, route-ordered

Moved verbatim from server.py. `openpyxl` is imported lazily inside the
handler (heavy dependency, only needed on export). `db` is lazy-imported from
`server` per request.

NOTE on the auth dependency: this endpoint depends on `server.get_current_user`
*directly* (not the usual local `_current_user` wrapper) because
`tests/test_xlsx_export_lock.py` overrides it via
`app.dependency_overrides[server.get_current_user]`. The dependency callable
must be that exact object for the override to take effect. `get_current_user`
is defined early in server.py (well before this module is imported at the
include-router block), so the module-level import is safe.
"""
from __future__ import annotations

import io
import logging

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response

from server import get_current_user  # noqa: E402 — defined before this module loads

logger = logging.getLogger("server")
router = APIRouter()


@router.get("/stops/export/xlsx")
async def export_stops_xlsx(current_user=Depends(get_current_user)):
    """Export all stops as an Excel (.xlsx) file, ordered by route sequence.

    The `#` column reflects the Sharpie-locked `original_sequence` whenever
    the route has been confirmed — that's the value the driver wrote on the
    physical box, so it's what the spreadsheet must show. Pre-confirm rows
    fall back to the live drive `order` (which may still re-shuffle on the
    next optimise). Sort key follows the same rule so the spreadsheet's
    row order ALWAYS matches the displayed `#` column, never the other
    way around.
    """
    from server import db  # noqa: WPS433
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    def _num(v, default=None):
        # Coerce DB value to float; raw Mongo docs bypass the Pydantic Stop
        # model, so lat/lng/weight can be None or string-typed on legacy /
        # failed-geocode rows. round(None, 6) raises TypeError → HTTP 500.
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            try:
                return float(v.strip())
            except ValueError:
                return default
        return default

    try:
        return await _export_stops_xlsx_inner(current_user, db, openpyxl, Font, PatternFill, Alignment, Border, Side, _num)
    except Exception:
        logger.exception("[export/xlsx] Failed to generate XLSX for user=%s", current_user.user_id)
        raise HTTPException(status_code=500, detail="Failed to generate export — please try again or contact support.")


async def _export_stops_xlsx_inner(current_user, db, openpyxl, Font, PatternFill, Alignment, Border, Side, _num):
    raw_stops = await db.stops.find({"user_id": current_user.user_id}, {"_id": 0}).to_list(2000)
    # Sort by (locked-sequence ?? live-order). Once original_sequence is
    # written by /api/routes/confirm it never moves, so the export stays
    # stable across re-optimisations of partially-confirmed routes.
    def _sort_key(s):
        seq = s.get("original_sequence")
        if isinstance(seq, int) and seq > 0:
            return (0, seq)
        ordr = s.get("order")
        return (1, ordr if isinstance(ordr, int) else 999_999)
    all_stops = sorted(raw_stops, key=_sort_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Route Stops"

    # Header style
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Group sibling parcels (same address, same `order` after sort) so
    # the spreadsheet reads as "stop → its parcels" instead of a flat list
    # where the driver has to scan addresses to spot multi-parcel stops.
    # Each ROW = one physical parcel with its own tracking number + weight
    # (the actual unit we ship). A subtotal row appears under each stop
    # whenever it contains 2+ parcels.
    headers = ["#", "Name", "Address", "Status", "Tracking #", "Weight (kg)", "Latitude", "Longitude", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    completed_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    subtotal_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    subtotal_font = Font(name="Calibri", italic=True, color="7F6000", size=10)
    total_weight = 0.0
    current_addr = None       # The address being grouped (siblings share this)
    addr_parcel_count = 0     # How many parcels seen so far for current_addr
    addr_weight_sum = 0.0     # Subtotal weight for current_addr
    excel_row = 2             # Pointer into worksheet (data rows + injected subtotals)

    def _emit_subtotal(rownum, parcel_count, wt_sum):
        """Inject a yellow subtotal banner below a multi-parcel stop."""
        if parcel_count < 2:
            return rownum  # Single-parcel stops don't need a subtotal banner
        for col in range(1, 5):
            cell = ws.cell(row=rownum, column=col, value="")
            cell.fill = subtotal_fill
            cell.border = thin_border
        label_cell = ws.cell(row=rownum, column=5,
                             value=f"{parcel_count} parcels — subtotal:")
        label_cell.fill = subtotal_fill
        label_cell.font = subtotal_font
        label_cell.alignment = Alignment(horizontal="right")
        label_cell.border = thin_border
        sub_cell = ws.cell(row=rownum, column=6, value=round(wt_sum, 2))
        sub_cell.fill = subtotal_fill
        sub_cell.font = subtotal_font
        sub_cell.alignment = Alignment(horizontal="right")
        sub_cell.border = thin_border
        for col in (7, 8, 9):
            cell = ws.cell(row=rownum, column=col, value="")
            cell.fill = subtotal_fill
            cell.border = thin_border
        return rownum + 1

    for stop in all_stops:
        stop_addr = stop.get("address", "")
        # Boundary detection — if this stop's address differs from the one
        # we were grouping, flush a subtotal row for the previous group.
        if current_addr is not None and stop_addr != current_addr:
            excel_row = _emit_subtotal(excel_row, addr_parcel_count, addr_weight_sum)
            addr_parcel_count = 0
            addr_weight_sum = 0.0
        current_addr = stop_addr
        addr_parcel_count += 1

        order_num = stop.get("order", excel_row - 2)
        is_completed = stop.get("completed", False)
        # Pin number rendered in the # column follows the same Sharpie-marker
        # contract as the map pins:
        #  • Locked: original_sequence (immutable post first /routes/confirm)
        #  • Tentative: order + 1 (server-stamped optimised drive position)
        # The driver cannot get a different number on the box vs the
        # spreadsheet vs the map — they are all bound to original_sequence
        # the moment the route is confirmed, never to the live `order`
        # which can shift on re-optimise.
        seq = stop.get("original_sequence")
        if isinstance(seq, int) and seq > 0:
            display_num = seq
        else:
            display_num = (order_num + 1) if isinstance(order_num, int) else (excel_row - 1)
        # Weight is optional — empty string when missing so the column reads
        # cleanly in Excel (vs `None` which renders literally as "None").
        # Sum only the populated values so the totals reflect ACTUAL known
        # load, not under-counted phantom zeros.
        raw_w = _num(stop.get("weight"))
        weight_val = round(raw_w, 2) if raw_w is not None else ""
        if isinstance(weight_val, float):
            total_weight += weight_val
            addr_weight_sum += weight_val
        lat = _num(stop.get("latitude"))
        lng = _num(stop.get("longitude"))
        values = [
            display_num,
            stop.get("name", ""),
            stop_addr,
            "Completed" if is_completed else "Pending",
            stop.get("tracking_number", "") or "",
            weight_val,
            round(lat, 6) if lat is not None else "",
            round(lng, 6) if lng is not None else "",
            stop.get("notes", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.border = thin_border
            if is_completed:
                cell.fill = completed_fill
        # Right-align the Weight column for readability — numbers sit
        # cleanly under the header when right-aligned.
        ws.cell(row=excel_row, column=6).alignment = Alignment(horizontal="right")
        # Tracking column also right-aligned + monospace-feel — barcode IDs
        # sort better visually that way.
        ws.cell(row=excel_row, column=5).alignment = Alignment(horizontal="left")
        excel_row += 1

    # Flush the final group's subtotal (if it was a multi-parcel stop).
    if current_addr is not None:
        excel_row = _emit_subtotal(excel_row, addr_parcel_count, addr_weight_sum)

    # Summary footer — total weight across all populated rows. Helps the
    # driver / dispatch sanity-check vehicle load capacity in one glance.
    if all_stops:
        footer_row = excel_row
        footer_font = Font(name="Calibri", bold=True, size=11)
        footer_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for col in range(1, 5):
            cell = ws.cell(row=footer_row, column=col,
                           value="Grand Total Weight" if col == 4 else "")
            cell.font = footer_font
            cell.fill = footer_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if col == 4 else "left")
        # Tracking column header in footer — leave blank
        blank = ws.cell(row=footer_row, column=5, value="")
        blank.font = footer_font
        blank.fill = footer_fill
        blank.border = thin_border
        total_cell = ws.cell(row=footer_row, column=6, value=round(total_weight, 2))
        total_cell.font = footer_font
        total_cell.fill = footer_fill
        total_cell.border = thin_border
        total_cell.alignment = Alignment(horizontal="right")
        # Trailing blank cells get the same fill so the row reads as a
        # single banded footer.
        for col in (7, 8, 9):
            cell = ws.cell(row=footer_row, column=col, value="")
            cell.fill = footer_fill
            cell.border = thin_border

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=route_stops.xlsx"},
    )
